"""
This module contains the API routes for exporting data in various formats.
"""

import csv
import datetime
import os
from enum import Enum
from io import BytesIO, StringIO
from typing import Optional

from bson import ObjectId
from fastapi import APIRouter, Header, HTTPException, Query, Response
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pytz import timezone  # type: ignore
from reportlab.lib import colors  # type: ignore
from reportlab.lib.pagesizes import letter  # type: ignore
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet  # type: ignore
from reportlab.lib.units import inch  # type: ignore
from reportlab.platypus import (  # type: ignore
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from api.utils.auth import verify_token
from config.config import MONGO_URI, TIME_ZONE

router = APIRouter(prefix="/exports", tags=["Exports"])

# MongoDB setup
client: AsyncIOMotorClient = AsyncIOMotorClient(MONGO_URI)
db = client.mmdb
expenses_collection = db.expenses
accounts_collection = db.accounts
users_collection = db.users


class ExportType(str, Enum):
    """Enum for export types."""
    EXPENSES = "expenses"
    ACCOUNTS = "accounts"
    CATEGORIES = "categories"


# Utility function to fetch data
async def fetch_data(
    user_id: str, from_date: Optional[datetime.date], to_date: Optional[datetime.date]
):
    """Fetch data from the database based on user ID and date range."""
    if from_date and to_date and from_date > to_date:
        raise HTTPException(
            status_code=422,
            detail="Invalid date range: 'from_date' must be before 'to_date'",
        )

    from_dt = (
        datetime.datetime.combine(from_date, datetime.time.min) if from_date else None
    )
    to_dt = datetime.datetime.combine(to_date, datetime.time.max) if to_date else None

    query = {"user_id": user_id}
    if from_dt and to_dt:
        query["date"] = {"$gte": from_dt, "$lte": to_dt}
    elif from_dt:
        query["date"] = {"$gte": from_dt}
    elif to_dt:
        query["date"] = {"$lte": to_dt}

    expenses = await expenses_collection.find(query).to_list(1000)
    accounts = await accounts_collection.find({"user_id": user_id}).to_list(100)
    user = await users_collection.find_one({"_id": ObjectId(user_id)})

    return expenses, accounts, user

def write_expenses_to_sheet(sheet: Worksheet, expenses: list):
    """Write expenses data to the given worksheet."""
    sheet.append(
        ["date", "amount", "currency", "category", "description", "account_name", "_id"]
    )
    for expense in expenses:
        sheet.append(
            [
                expense["date"].isoformat() if expense.get("date") else "",
                expense["amount"],
                expense["currency"],
                expense["category"],
                expense.get("description", ""),
                expense["account_name"],
                str(expense["_id"]),
            ]
        )

def write_accounts_to_sheet(sheet: Worksheet, accounts: list):
    """Write accounts data to the given worksheet."""
    sheet.append(["name", "balance", "currency", "_id"])
    for account in accounts:
        sheet.append(
            [
                account["name"],
                account["balance"],
                account["currency"],
                str(account["_id"]),
            ]
        )

def write_categories_to_sheet(sheet: Worksheet, categories: dict):
    """Write categories data to the given worksheet."""
    sheet.append(["name", "monthly_budget"])
    for category_name, category_data in categories.items():
        sheet.append([category_name, category_data["monthly_budget"]])

@router.get("/xlsx")
async def data_to_xlsx(
    token: str = Header(None),
    from_date: Optional[datetime.date] = Query(None),
    to_date: Optional[datetime.date] = Query(None),
) -> Response:
    """
    Export all expenses, accounts, and categories for a user to an XLSX file.

    Args:
        token (str): Authentication token.

    Returns:
        Response: XLSX file containing expenses, accounts, and categories data.
    """
    user_id = await verify_token(token)
    expenses, accounts, user = await fetch_data(user_id, from_date, to_date)

    if not expenses and not accounts and not user:
        raise HTTPException(status_code=404, detail="No data found")

    workbook = Workbook()

    # Write expenses
    expenses_sheet: Optional[Worksheet] = workbook.active
    if expenses_sheet is not None:
        expenses_sheet.title = "Expenses"
        write_expenses_to_sheet(expenses_sheet, expenses)

    # Write accounts
    accounts_sheet: Optional[Worksheet] = workbook.create_sheet(title="Accounts")
    if accounts_sheet is not None:
        write_accounts_to_sheet(accounts_sheet, accounts)

    # Write categories
    categories_sheet: Optional[Worksheet] = workbook.create_sheet(title="Categories")
    if categories_sheet is not None and user and "categories" in user:
        write_categories_to_sheet(categories_sheet, user["categories"])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = "attachment; filename=data.xlsx"
    return response

def create_paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    """Create a paragraph with the given text and style."""
    return Paragraph(text, style)

def create_table(data: list, col_widths: list, styles: TableStyle) -> Table:
    """Create a table with the given data, column widths, and styles."""
    table = Table(data, colWidths=col_widths)
    table.setStyle(styles)
    return table

@router.get("/csv")
async def data_to_csv(
    token: str = Header(None),
    export_type: ExportType = Query(...),
    from_date: Optional[datetime.date] = Query(None),
    to_date: Optional[datetime.date] = Query(None),
) -> Response:
    """
    Export expenses, accounts, or categories for a user to a CSV file.

    Args:
        token (str): Authentication token.
        export_type (ExportType): Type of data to export (expenses, accounts, categories).

    Returns:
        Response: CSV file containing the selected data.
    """
    user_id = await verify_token(token)
    expenses, accounts, user = await fetch_data(user_id, from_date, to_date)
    output = StringIO()
    writer = csv.writer(output)

    if export_type == ExportType.EXPENSES:
        if not expenses:
            raise HTTPException(status_code=404, detail="No expenses found")
        writer.writerow(
            [
                "date",
                "amount",
                "currency",
                "category",
                "description",
                "account_name",
                "_id",
            ]
        )
        for expense in expenses:
            writer.writerow(
                [
                    expense["date"].isoformat() if expense.get("date") else "",
                    expense["amount"],
                    expense["currency"],
                    expense["category"],
                    expense.get("description", ""),
                    expense["account_name"],
                    str(expense["_id"]),
                ]
            )
    elif export_type == ExportType.ACCOUNTS:
        if not accounts:
            raise HTTPException(status_code=404, detail="No accounts found")
        writer.writerow(["name", "balance", "currency", "_id"])
        for account in accounts:
            writer.writerow(
                [
                    account["name"],
                    account["balance"],
                    account["currency"],
                    str(account["_id"]),
                ]
            )
    elif export_type == ExportType.CATEGORIES:
        if not user or "categories" not in user:
            raise HTTPException(status_code=404, detail="No categories found")
        writer.writerow(["name", "monthly_budget"])
        for category_name, category_data in user["categories"].items():
            writer.writerow([category_name, category_data["monthly_budget"]])

    response = Response(content=output.getvalue(), media_type="text/csv")
    response.headers[
        "Content-Disposition"
    ] = f"attachment; filename={export_type.value}.csv"
    return response


@router.get("/pdf")
async def data_to_pdf(
    token: str = Header(None),
    from_date: Optional[datetime.date] = Query(None),
    to_date: Optional[datetime.date] = Query(None),
) -> Response:
    """
    Export all expenses, accounts, and categories for a user to a PDF file within a date range.

    Args:
        token (str): Authentication token.
        from_date (datetime.date, optional): Start date for filtering expenses (inclusive).
        to_date (datetime.date, optional): End date for filtering expenses (inclusive).

    Returns:
        Response: PDF file containing expenses, accounts, and categories data.
    """
    # pylint: disable=too-many-locals, too-many-statements
    user_id = await verify_token(token)
    expenses, accounts, user = await fetch_data(user_id, from_date, to_date)

    if not expenses and not accounts and not user:
        raise HTTPException(status_code=404, detail="No data found")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        title=f"MM PDF Export - {user['username']}",
        lang="en-gb",
    )
    styles = getSampleStyleSheet()
    elements = []

    # Define a custom style for the title
    title_style = ParagraphStyle(
        name="Title",
        parent=styles["Title"],
        fontSize=32,  # Increase font size
        spaceAfter=12,
    )

    # Define a custom style for the centered description
    centered_style = ParagraphStyle(
        name="Centered",
        parent=styles["Normal"],
        alignment=1,  # Center alignment
    )

    # Add heading, logo, application description, and TOC on the first page
    elements.append(create_paragraph("MONEY MANAGER", title_style))
    elements.append(Spacer(1, 12))
    logo_path = os.path.abspath(
        os.path.join(os.path.dirname(__file__), "../../docs/logo/logo.png")
    )
    logo = Image(logo_path)
    logo.drawHeight = 3.0 * inch * logo.drawHeight / logo.drawWidth
    logo.drawWidth = 3.0 * inch
    elements.append(logo)
    elements.append(Spacer(1, 18))
    app_description = """
    <b>Money Manager</b> is a comprehensive financial management tool designed to help you track your expenses, manage your accounts, and set budgets for various categories.
    With our application, you can easily export your financial data in various formats including XLSX, CSV, and PDF.
    """
    elements.append(create_paragraph(app_description, centered_style))
    elements.append(Spacer(1, 18))
    elements.append(create_paragraph(f"PDF Report for - {user['username']}", styles["Title"]))
    elements.append(Spacer(1, 36))

    # Table of Contents
    toc = [
        create_paragraph("<link href='#expenses'>1. Expenses</link>", styles["Normal"]),
        create_paragraph("<link href='#accounts'>2. Accounts</link>", styles["Normal"]),
        create_paragraph("<link href='#categories'>3. Categories</link>", styles["Normal"]),
    ]
    elements.append(create_paragraph("Table of Contents", styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.extend(toc)
    elements.append(PageBreak())

    # Helper function to wrap text in table cells
    def wrap_text(data):
        wrapped_data = []
        for row in data:
            wrapped_row = []
            for cell in row:
                wrapped_row.append(Paragraph(str(cell), styles["Normal"]))
            wrapped_data.append(wrapped_row)
        return wrapped_data

    # Expenses
    elements.append(create_paragraph("<a name='expenses'/>Expenses", styles["Title"]))
    elements.append(Spacer(1, 12))
    if from_date and to_date:
        if from_date == to_date:
            date_range_text = f"Date: {from_date}"
        else:
            date_range_text = f"Date Range: {from_date} to {to_date}"
    elif from_date:
        date_range_text = f"Date Range: From {from_date}"
    elif to_date:
        date_range_text = f"Date Range: To {to_date}"
    else:
        date_range_text = "Date Range: All"
    elements.append(create_paragraph(date_range_text, styles["Normal"]))
    elements.append(Spacer(1, 12))
    expenses_data = [
        ["Date", "Amount", "Currency", "Category", "Description", "Account Name", "ID"]
    ]
    for expense in expenses:
        expenses_data.append(
            [
                expense["date"].isoformat() if expense.get("date") else "",
                expense["amount"],
                expense["currency"],
                expense["category"],
                expense.get("description", ""),
                expense["account_name"],
                str(expense["_id"]),
            ]
        )
    expenses_table = create_table(
        wrap_text(expenses_data), [60, 60, 60, 60, 120, 80, 80],
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    elements.append(expenses_table)
    elements.append(PageBreak())

    # Accounts
    elements.append(create_paragraph("<a name='accounts'/>Accounts", styles["Title"]))
    elements.append(Spacer(1, 12))
    accounts_data = [["Name", "Balance", "Currency", "ID"]]
    for account in accounts:
        accounts_data.append(
            [
                account["name"],
                account["balance"],
                account["currency"],
                str(account["_id"]),
            ]
        )
    accounts_table = create_table(
        wrap_text(accounts_data), [100, 100, 100, 100],
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    elements.append(accounts_table)
    elements.append(PageBreak())

    # Categories
    elements.append(create_paragraph("<a name='categories'/>Categories", styles["Title"]))
    elements.append(Spacer(1, 12))
    categories_data = [["Name", "Monthly Budget"]]
    if user and "categories" in user:
        for category_name, category_data in user["categories"].items():
            categories_data.append([category_name, category_data["monthly_budget"]])
    categories_table = create_table(
        wrap_text(categories_data), [200, 200],
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    elements.append(categories_table)

    # Footer with date of export, "Money Manager V2", and page number
    def footer(canvas, doc):
        """Footer with date of export, 'Money Manager V2', and page number."""
        canvas.saveState()
        tz = timezone(TIME_ZONE)
        footer_text = (
            f"Money Manager V2 - Exported on "
            f"{datetime.datetime.now(tz).strftime('%Y-%m-%d %H:%M:%S')}"
        )
        canvas.setFont("Helvetica", 9)
        canvas.drawString(inch, 0.75 * inch, footer_text)
        canvas.drawRightString(7.5 * inch, 0.75 * inch, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    buffer.seek(0)

    response = Response(content=buffer.getvalue(), media_type="application/pdf")
    response.headers["Content-Disposition"] = "attachment; filename=data.pdf"
    return response
